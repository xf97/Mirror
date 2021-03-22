from os.path import isfile
from openpyxl import load_workbook



class ExcelToDict:
    """
    将Excel文件对象转成Python字典对象
    """

    def __init__(self, file_dir, title_row=0):
        # 工作簿文件的路径
        self.file_dir = file_dir
        # 标题行位于第几行
        self.title_row = int(title_row)
        self.data_dict = {}
        self.work_book = None

    def open_object(self):
        """打开工作簿对象"""
        valid = isfile(self.file_dir)
        # file_dir指向的文件是否不存在
        if not valid:
            raise Exception('文件路径 {0} 不存在'.format(self.file_dir))
        self.work_book = load_workbook(filename=self.file_dir)

    def read_excel(self):
        """读取工作簿数据"""
        if not self.work_book:
            raise Exception('需要先调用 open_object() 方法以打开工作簿对象')
        for sheet_name in self.work_book.sheetnames:
            # 每个工作表的字典
            data_dict_sheet = {'title_row': [], 'value_row': {}}
            # 获取工作表对象
            ws = self.work_book[sheet_name]
            # 预先创建工作表中每一行的字典
            for i in range(ws.max_row - 1 - self.title_row):
                data_dict_sheet['value_row'][i] = {}
            # 遍历所有列
            columns = tuple(ws.columns)
            for column in columns:
                # 每一列的标题
                title = column[self.title_row].value
                # 记录每列的标题
                data_dict_sheet['title_row'].append(title)
                row_num = 0
                # 遍历每一列中的所有值
                for col in column:
                    # 忽略每一列的标题行
                    if column.index(col) <= self.title_row:
                        continue
                    data_dict_sheet['value_row'][row_num][title] = col.value
                    row_num += 1
            # 记录每个工作表的数据字段
            self.data_dict[sheet_name] = data_dict_sheet

    def check(self, check_item=None, sheet_name=None, sheet_index=0):
        """
        在所选工作表中校验是否包含业务需要的所有标题名称
        :param check_item: 所选工作表中需要校验的标题列表
        :param sheet_name: 以名称形式选择工作表（优先）
        :param sheet_index: 以下标形式选择工作表
        :return: {'result': True, 'exception': None}
        """
        if not self.data_dict:
            return {'result': False, 'exception': '需要先调用 read_excel() 方法以读取工作簿数据'}
        if check_item is None:
            check_item = []
        if sheet_name:
            if sheet_name not in self.data_dict:
                return {'result': False, 'exception': '不存在名为 {0} 的工作表'.format(sheet_name)}
            # 直接获得对应的工作表数据
            data_sheet = self.data_dict[sheet_name]
        else:
            # 通过下标获取对应的工作表名称
            data_dict_keys = tuple(self.data_dict.keys())
            if len(data_dict_keys) <= int(sheet_index):
                return {'result': False, 'exception': '不存在下标为 {0} 的工作表'.format(sheet_index)}
            _sheet_name = data_dict_keys[int(sheet_index)]
            # 间接获得对应的工作表数据
            data_sheet = self.data_dict[_sheet_name]
        # 判断工作表中是否包含业务需要的所有标题
        if not set(check_item).issubset(set(data_sheet['title_row'])):
            return {'result': False, 'exception': '工作表中未包含业务需要的 {0} 标题'.format(check_item)}
        return {'result': True, 'exception': None}


if __name__ == '__main__':
    excel_to_dict = ExcelToDict('..\\data\\dataForExp.xlsx')
    excel_to_dict.open_object()
    print('工作簿对象：', excel_to_dict.work_book)
    excel_to_dict.read_excel()
    print('工作簿数据：', excel_to_dict.data_dict)
    print('工作簿校验（异常演示）：', excel_to_dict.check(['标题四']))
    print('工作簿校验（正常演示）：', excel_to_dict.check(['标题一', '标题二']))